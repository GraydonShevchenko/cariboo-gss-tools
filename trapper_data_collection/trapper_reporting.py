import sys, os
import pandas as pd
import boto3
import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from arcgis.gis import GIS
from argparse import ArgumentParser
import logging

from util.environment import Environment

import trap_config


def run_app():
    ago_user, ago_pass, obj_store_user, obj_store_secret, obj_store_host, logger = get_input_parameters()
    report = TrapReport(ago_user=ago_user, ago_pass=ago_pass, obj_store_user=obj_store_user, 
                       obj_store_secret=obj_store_secret, obj_store_host=obj_store_host, logger=logger)
    
    report.download_attachments()
    report.create_excel()

    del report


def get_input_parameters():
    """
    Function:
        Sets up parameters and the logger object
    Returns:
        tuple: user entered parameters required for tool execution
    """
    try:
        parser = ArgumentParser(description='This script is used to update the Traps AGOL feature layer based on information entered in the trap check table')
        # parser.add_argument('ago_user', nargs='?', type=str, help='AGOL Username')
        # parser.add_argument('ago_pass', nargs='?', type=str, help='AGOL Password')
        parser.add_argument('--log_level', default='INFO', choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'],
                            help='Log level')
        parser.add_argument('--log_dir', help='Path to log directory')

        args = parser.parse_args()
        try:
            ago_user = trap_config.AGO_USER
            ago_pass = trap_config.AGO_PASS
            obj_store_user = trap_config.OBJ_STORE_USER
            obj_store_secret = trap_config.OBJ_STORE_SECRET
            obj_store_host = trap_config.OBJ_STORE_HOST
        except:
            ago_user = os.environ['AGO_USER']
            ago_pass = os.environ['AGO_PASS']
            obj_store_user = os.environ['OBJ_STORE_USER']
            obj_store_secret = os.environ['OBJ_STORE_SECRET']
            obj_store_host = os.environ['OBJ_STORE_HOST']

        logger = Environment.setup_logger(args)

        return ago_user, ago_pass, obj_store_user, obj_store_secret, obj_store_host, logger

    except Exception as e:
        logging.error('Unexpected exception. Program terminating: {}'.format(e.message))
        raise Exception('Errors exist')


class TrapReport:
    def __init__(self, ago_user, ago_pass, obj_store_user, obj_store_secret, obj_store_host, logger) -> None:
        self.ago_user = ago_user
        self.ago_pass = ago_pass
        self.obj_store_user = obj_store_user
        self.obj_store_secret = obj_store_secret
        self.obj_store_host = obj_store_host
        self.logger = logger

        self.portal_url = trap_config.MAPHUB
        self.ago_traps = trap_config.TRAPS
        self.ago_fisher = trap_config.FISHER

        self.trapper_bucket = 'rcbgss'
        self.bucket_prefix ='trapper_data_collection'

        self.logger.info('Connecting to map hub')
        self.gis = GIS(url=self.portal_url, username=self.ago_user, password=self.ago_pass, expiration=9999)
        self.logger.info('Connection successful')

        self.logger.info('Connecting to object storage')
        self.boto_resource = boto3.resource(service_name='s3', 
                                            aws_access_key_id=self.obj_store_user,
                                            aws_secret_access_key=self.obj_store_secret, 
                                            endpoint_url=f'https://{self.obj_store_host}')

    def __del__(self) -> None:
        self.logger.info('Disconnecting from maphub')
        del self.gis
        self.logger.info('Closing object storage connection')
        del self.boto_resource

    def list_contents(self) -> list:
        obj_bucket = self.boto_resource.Bucket(self.trapper_bucket)
        lst_objects = []
        for obj in obj_bucket.objects.all():
            lst_objects.append(os.path.basename(obj.key))

        return lst_objects


    def download_attachments(self) -> None:
        """
        Function:
            Master function to download attachments for all required layers in arcgis online
        Returns:
            None
        """
        lst_pictures = self.list_contents()

        self.copy_to_object_storage(ago_layer=self.ago_traps, layer_name='traps', 
                                    fld_picture='PICTURE', lst_os_pictures=lst_pictures, folder='trap_setup')
        
        self.copy_to_object_storage(ago_layer=self.ago_traps, layer_name='trap checks', 
                                    fld_picture='PICTURE', lst_os_pictures=lst_pictures, folder='trap_check')
        
        self.copy_to_object_storage(ago_layer=self.ago_fisher, layer_name='fisher', 
                                    fld_picture='PICTURE', lst_os_pictures=lst_pictures, folder='fisher')
        

    def copy_to_object_storage(self, ago_layer, layer_name, fld_picture, lst_os_pictures, folder) -> None:
        """
        Function:
            Function used to download attachments from arcgis online layers and copy them to object storage.
        Returns:
            None
        """
        self.logger.info(f'Downloading photos on the {layer_name} layer')
        ago_item = self.gis.content.get(ago_layer)
        if layer_name != 'trap checks':
            ago_flayer = ago_item.layers[0]
        else:
            ago_flayer = ago_item.tables[0]

        ago_fset = ago_flayer.query()
        all_features = ago_fset.features
        if len(all_features) == 0:
            return

        lst_oids = ago_fset.sdf['OBJECTID'].tolist()

        for oid in lst_oids:
            lst_attachments = ago_flayer.attachments.get_list(oid=oid)
            if lst_attachments:
                original_feature = [f for f in all_features if f.attributes['OBJECTID'] == oid][0]
                lst_pictures = original_feature.attributes[fld_picture].split(',')
                lst_new_pictures = [pic for pic in lst_pictures if pic not in lst_os_pictures]
                if not lst_new_pictures:
                    continue

                for attach in lst_attachments:
                    attach_name = attach['name']
                    if attach_name in lst_new_pictures:
                        self.logger.info(f'Copying {attach_name} to object storage')
                        attach_id = attach['id']
                        attach_file = ago_flayer.attachments.download(oid=oid, attachment_id=attach_id)[0]
                        ostore_path = f'{self.bucket_prefix}/{folder}/{attach_name}'

                        self.boto_resource.meta.client.upload_file(attach_file, self.trapper_bucket, ostore_path)


    def create_excel(self) -> None:
        self.logger.info('Creating report')
        
        xl_report = 'trapper_data_report.xlsx'
        with pd.ExcelWriter(xl_report, date_format='yyyy-mm-dd', datetime_format='yyyy-mm-dd') as xl_writer:
            self.create_sheet(xl_writer=xl_writer, sheet_name='traps', ago_layer=self.ago_traps, 
                              drop_columns=['GlobalID', 'OBJECTID', 'EDIT_DATE', 'CALCULATE_DATE', 'SHAPE'], date_field='START_DATE')
            
            self.create_sheet(xl_writer=xl_writer, sheet_name='trap checks', ago_layer=self.ago_traps, 
                              drop_columns=['GlobalID', 'OBJECTID', 'EDIT_DATE', 'CALCULATE_DATE', 'TRAPSET_TYPES'], date_field='CHECK_DATE')
            
            self.create_sheet(xl_writer=xl_writer, sheet_name='fisher', ago_layer=self.ago_fisher, 
                              drop_columns=['GlobalID', 'OBJECTID', 'EDIT_DATE', 'CALCULATE_DATE', 'SHAPE'], date_field='OBSERVATION_DATE')

        ostore_path = f'{self.bucket_prefix}/{os.path.basename(xl_report)}'

        self.logger.info('Uploading document to object storage')
        self.boto_resource.meta.client.upload_file(xl_report, self.trapper_bucket, ostore_path)
    

    def create_sheet(self, xl_writer, sheet_name, ago_layer, drop_columns, date_field) -> None:
        self.logger.info(f'Generating {sheet_name} sheet')
        ago_item = self.gis.content.get(ago_layer)
        if sheet_name != 'trap checks':
            ago_flayer = ago_item.layers[0]
        else:
            ago_flayer = ago_item.tables[0]
        ago_fset = ago_flayer.query()
        if len(ago_fset.features) == 0:
            return
        df = ago_fset.sdf
        df.drop(drop_columns, axis=1, inplace=True)
        df[date_field] = pd.to_datetime(df[date_field]).dt.date
        df.to_excel(xl_writer, sheet_name=sheet_name, index=False)

        ws = xl_writer.sheets[sheet_name]

        dim_holder = DimensionHolder(worksheet=ws)

        for col in range(ws.min_column, ws.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)

            ws.column_dimensions = dim_holder
        
        for row in ws.iter_rows():
            for cell in row:
                cell.style.alignment.wrap_text=True


def get_col_widths(dataframe):
    # First we find the maximum length of the index column
    idx_max = max([len(str(s)) for s in dataframe.index.values] + [len(str(dataframe.index.name))])
    # Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right

    return [idx_max] + [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) + 4 for col in dataframe.columns]

if __name__ == '__main__':
    run_app()
